from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook

import time


PATH = 'venv\msedgedriver.exe'
user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36 Edg/107.0.1418.62'
edge_options = Options()
edge_options.add_experimental_option('detach',True)
edge_options.add_argument(f'user_agent={user_agent}')
edge_options.add_experimental_option("excludeSwitches", ["enable-logging"])
edge_options.add_argument("--disable-popup-blocking")
edge_options.add_argument('--disable-default-apps')
edge_options.add_argument('--allow-silent-push')
edge_options.add_argument('--disable-notifications')
edge_options.add_argument('--suppress-message-center-popups')
edge_options.add_argument('--inprivate')

edge_service = Service(PATH)
driver = webdriver.Edge(service=edge_service,options=edge_options)
driver.get('https://www.linkedin.com/')

#username = "bizera9722@gmail.com"
#password = "Poop123P"
username = "milkycake05@gmail.com"
password = "invicta"
driver.implicitly_wait(10)
driver.find_element(By.ID,'session_key').send_keys(username)
driver.find_element(By.ID,'session_password').send_keys(password,Keys.RETURN)

time.sleep(10)


#driver.get('https://www.linkedin.com/search/results/companies/?keywords=hemp%20gummies&origin=GLOBAL_SEARCH_HEADER&sid=J%40L')
# driver.get('https://www.linkedin.com/search/results/companies/?keywords=hemp%20gummies&origin=SWITCH_SEARCH_VERTICAL&page=9&searchId=fd114a8e-a721-478f-ae61-d53ea00922c9&sid=H)z')
#driver.get('https://www.linkedin.com/search/results/companies/?keywords=hemp%20gummies&origin=SWITCH_SEARCH_VERTICAL&page=29&searchId=fd114a8e-a721-478f-ae61-d53ea00922c9&sid=AOU')
#driver.get('https://www.linkedin.com/search/results/companies/?companyHqGeo=%5B%22103644278%22%5D&keywords=casino&origin=FACETED_SEARCH&sid=OpM')
driver.get('https://www.linkedin.com/search/results/companies/?companyHqGeo=%5B%22103644278%22%5D&keywords=casino&origin=FACETED_SEARCH&page=17&sid=4TE')


action_chains = ActionChains(driver)


try:
    #workbook = load_workbook('cbd companies.xlsx')
    workbook = load_workbook('casinos.xlsx')
    sheet = workbook.active
    last_row = sheet.max_row
    while(True):
        posts = driver.find_elements(By.XPATH,'.//ul//a[@class="app-aware-link "]')
        print(len(posts))
        #time.sleep(2)

        for post in posts:
                driver.switch_to.window(driver.window_handles[0])
                
                last_row+=1
                
                data = []
                driver.switch_to.window(driver.window_handles[0])
                print(post.get_attribute('innerText'))

                data.append(post.get_attribute('innerText'))

                action_chains.key_down(Keys.CONTROL).click(post).key_up(Keys.CONTROL).perform()

                try: 
                    driver.switch_to.window(driver.window_handles[1])
                except:
                    continue
                #time.sleep(2)
                driver.execute_script("window.scrollTo(0, 0);")

                driver.find_element(By.XPATH,'.//ul[@class="org-page-navigation__items "]//li[2]').click()
                             
                #time.sleep(1)
                try:
                    try:
                         
                        data.append(driver.find_element(By.CSS_SELECTOR,'p.break-words.white-space-pre-wrap.t-black--light.text-body-medium').get_attribute("innerText"))
                    except:
                         driver.close()
                         continue
                    
                    all_contact = driver.find_elements(By.XPATH,'//span[@class="link-without-visited-state"]')


                    for contact in all_contact:
                        contact_info = contact.get_attribute('innerText')
                        data.append(contact_info)
                        print(contact_info)
                        
                    driver.close()
                    #time.sleep(0.5)
                    for i, value in enumerate(data, start=1):
                        sheet.cell(row=last_row, column=i).value = value
                    #workbook.save('cbd companies.xlsx')
                    workbook.save('casinos.xlsx')

                finally:
                    pass
                #time.sleep(1)
                i+=1
                
        driver.switch_to.window(driver.window_handles[0])
        driver.find_element(By.XPATH,'//button[@aria-label="Next"]').click()
        #time.sleep(2)


finally:
      pass